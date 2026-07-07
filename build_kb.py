#!/usr/bin/env python3
"""
build_kb.py — Generador del knowledge base de Dr's Choice

Lee todos los PDFs de /pdfs/ y las URLs de urls.txt,
extrae texto y tablas estructuradas, y genera knowledge_base.json.

Uso manual:
    pip install pdfplumber requests beautifulsoup4
    python build_kb.py

Se ejecuta automáticamente vía GitHub Actions en cada push
que modifique /pdfs/ o urls.txt.
"""

import json
import os
import re
import requests
import unicodedata
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    import pdfplumber
except ImportError:
    print("⚠️  pdfplumber no instalado. Ejecuta: pip install pdfplumber")
    raise

# ── Configuración ─────────────────────────────────────────────
PDFS_DIR    = Path("pdfs")
URLS_FILE   = Path("urls.txt")
OUTPUT_FILE = Path("knowledge_base.json")
KB_BASE     = Path("brandbook.json")  # datos estáticos de empresa/brandbook
DATA_DIR    = Path("data")

# ── 1. Cargar datos base (empresa, brandbook, servicios) ──────

def cargar_base() -> dict:
    """
    Carga el JSON base con datos estáticos del brandbook:
    empresa, áreas, segmentos, propuesta de valor, etc.
    Estos datos NO vienen de los PDFs — se mantienen a mano.
    """
    if KB_BASE.exists():
        with open(KB_BASE, encoding="utf-8") as f:
            return json.load(f)
    # Si no existe la base, retornamos estructura mínima
    print("⚠️  brandbook.json no encontrado — usando estructura mínima")
    return {
        "empresa": {
            "nombre": "Dr's Choice",
            "fundacion": 1992,
            "slogan": "Nos mueve tu bienestar",
            "whatsapp": "+56 9 6159 8525",
            "email": "serviciocliente@doctorchoice.cl",
            "web": "https://drchoice.cl",
            "direccion": "Miguel Claro 954, Providencia, Santiago"
        },
        "areas": {},
        "segmentos": {},
        "propuesta_de_valor": {},
        "chatbot": {},
        "mensajes_clave": {},
        "keywords_rag": [],
        "intents": {},
        "variables_crm": {},
        "servicios": []
    }

# ── 2. Extracción de PDFs ─────────────────────────────────────

def extraer_tablas_pdf(page) -> list[dict]:
    """
    Extrae tablas de una página PDF y las convierte en
    lista de dicts con los headers como claves.
    """
    tablas = []
    for tabla in page.extract_tables():
        if not tabla or len(tabla) < 2:
            continue
        headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
                   for i, h in enumerate(tabla[0])]
        for fila in tabla[1:]:
            if any(c for c in fila):  # descartamos filas vacías
                row = {headers[i]: str(v).strip() if v else ""
                       for i, v in enumerate(fila)}
                tablas.append(row)
    return tablas


def limpiar_texto(texto: str) -> str:
    """Limpia el texto extraído de ruido tipográfico."""
    if not texto:
        return ""
    # Elimina saltos de línea múltiples
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    # Elimina espacios múltiples
    texto = re.sub(r" {2,}", " ", texto)
    # Elimina líneas con solo números de página o guiones
    texto = re.sub(r"^\s*[\d\-–—]+\s*$", "", texto, flags=re.MULTILINE)
    return texto.strip()


def detectar_catalogo(nombre_archivo: str) -> str:
    """Detecta el tipo de catálogo según el nombre del archivo."""
    nombre = nombre_archivo.lower()
    if "ortesis" in nombre or "órtesis" in nombre:
        return "ortesis"
    elif "rehabilitacion" in nombre or "rehabilitación" in nombre or "insumos" in nombre:
        return "rehabilitacion"
    elif "tecnolog" in nombre:
        return "tecnologias"
    return "general"


def procesar_pdf(pdf_path: Path) -> dict:
    """
    Extrae texto y tablas de un PDF.
    Retorna dict con:
    - texto_completo: todo el texto del PDF
    - tablas: todas las tablas extraídas
    - paginas: número de páginas
    - catalogo: tipo detectado
    - fuente: nombre del archivo
    """
    catalogo = detectar_catalogo(pdf_path.name)
    texto_paginas = []
    todas_tablas = []

    print(f"  📄 Procesando {pdf_path.name}...")

    with pdfplumber.open(pdf_path) as pdf:
        n_paginas = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            # Texto
            texto = page.extract_text(x_tolerance=3, y_tolerance=3)
            if texto:
                texto_paginas.append(limpiar_texto(texto))

            # Tablas
            tablas = extraer_tablas_pdf(page)
            for tabla in tablas:
                tabla["_pagina"] = i + 1
                tabla["_catalogo"] = catalogo
                todas_tablas.append(tabla)

    texto_completo = "\n\n".join(texto_paginas)
    print(f"    ✅ {n_paginas} páginas | {len(texto_completo)} chars | {len(todas_tablas)} filas de tablas")

    return {
        "fuente": pdf_path.name,
        "catalogo": catalogo,
        "paginas": n_paginas,
        "texto_completo": texto_completo,
        "tablas": todas_tablas
    }


def extraer_productos_de_texto(texto: str, catalogo: str) -> list[dict]:
    """
    Usa heurísticas para identificar productos en el texto extraído.
    Busca patrones como SKU, nombres de productos, descripciones.
    
    Para catálogos bien estructurados esto funciona bien.
    Para PDFs con layouts complejos, puede requerir ajuste manual.
    """
    productos = []
    
    # Patrón básico: busca bloques que parezcan fichas de producto
    # Ajustar según la estructura real de los PDFs de Dr's Choice
    bloques = re.split(r"\n(?=[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-záéíóúñ\s]{5,})\n", texto)
    
    for bloque in bloques:
        bloque = bloque.strip()
        if len(bloque) < 30:
            continue
            
        # Busca SKU (patrones comunes en catálogos médicos)
        sku_match = re.search(
            r"(?:SKU|Ref\.?|Código|Code)[:\s]+([A-Z0-9\-\.\/]+)",
            bloque, re.IGNORECASE
        )
        
        if sku_match:
            sku = sku_match.group(1).strip()
            # Primera línea como nombre del producto
            lineas = [l.strip() for l in bloque.split("\n") if l.strip()]
            nombre = lineas[0] if lineas else "Producto sin nombre"
            descripcion = " ".join(lineas[1:5]) if len(lineas) > 1 else ""
            
            productos.append({
                "nombre": nombre,
                "sku": sku,
                "catalogo": catalogo,
                "categoria": "",
                "marca": None,
                "descripcion": limpiar_texto(descripcion),
                "indicaciones": [],
                "stock": "consultar",
                "_fuente_auto": True  # marcamos como extraído automáticamente
            })
    
    return productos

# ── 3. Extracción de Excel tabular ─────────────────────────────

def normalizar_clave(texto: str) -> str:
    """
    Normaliza headers y nombres para comparar sin acentos, mayúsculas ni símbolos.
    """
    texto = str(texto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def limpiar_celda(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor)).strip()
    return str(valor).strip()


def limpiar_precio(valor):
    """
    Convierte precios del Excel a entero CLP cuando sea posible.
    """
    if valor is None or valor == "":
        return None

    if isinstance(valor, (int, float)):
        return int(valor)

    texto = str(valor).strip()
    texto = re.sub(r"[^\d,\.]", "", texto)

    if not texto:
        return None

    # Formato chileno frecuente: 1.299.990
    if "," not in texto:
        texto = texto.replace(".", "")
    else:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return int(float(texto))
    except ValueError:
        return None


def split_indicaciones(texto: str) -> list[str]:
    """
    Convierte aplicaciones terapéuticas en lista breve para el RAG.
    """
    texto = limpiar_celda(texto)
    if not texto:
        return []

    partes = re.split(r"[;,]", texto)
    return [p.strip() for p in partes if p.strip()]


def detectar_fila_header_excel(ws) -> tuple[int, list[str]]:
    """
    El Excel tiene una nota en fila 1 y los headers reales en fila 2.
    Esta función detecta la fila que contiene Nombre y Marca.
    """
    alias = {
        "sku": "sku",
        "precio_rreferencia_base_neto": "precio_referencia_neto",
        "precio_referencia_base_neto": "precio_referencia_neto",
        "nombre": "nombre",
        "marca": "marca",
        "descripicion": "descripcion",
        "descripcion": "descripcion",
        "escalable": "escalable",
        "aplicaciones_terapias": "aplicaciones_terapias",
        "especificaciones_tecnicas_generales": "especificaciones_tecnicas",
        "vertical": "vertical",
        "pais_de_origen": "pais_origen",
        "proveedor": "proveedor",
        "perfil_comprador_ideal": "perfil_comprador_ideal",
        "tier": "tier",
        "url_web": "url_web",
        "se_puede_comprar_en_tienda_online": "canal_tienda_online",
        "url_tol": "url_tol",
    }

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers_raw = [normalizar_clave(c) for c in row]
        if "nombre" in headers_raw and "marca" in headers_raw:
            headers = [alias.get(h, h) for h in headers_raw]
            return idx, headers

    raise ValueError("No se encontró fila de headers en el Excel")


def sku_oficial_o_vacio(sku_ref: str) -> str:
    """
    En este Excel los SKU son referenciales.
    Si son solo números, no los usamos como SKU oficial.
    """
    sku_ref = limpiar_celda(sku_ref)
    if not sku_ref:
        return ""

    # Los SKU numéricos del Excel parecen IDs referenciales, no SKU reales.
    if sku_ref.isdigit():
        return ""

    return sku_ref


def producto_desde_excel(row: dict, fila_excel: int, archivo: str) -> dict | None:
    nombre = limpiar_celda(row.get("nombre"))
    if not nombre:
        return None

    sku_ref = limpiar_celda(row.get("sku"))
    sku = sku_oficial_o_vacio(sku_ref)

    descripcion = limpiar_celda(row.get("descripcion"))
    aplicaciones = limpiar_celda(row.get("aplicaciones_terapias"))
    especificaciones = limpiar_celda(row.get("especificaciones_tecnicas"))

    url_tol = limpiar_celda(row.get("url_tol"))
    if url_tol.lower().startswith("http"):
        url_tienda_online = url_tol
        titulo_tienda_online = ""
    else:
        url_tienda_online = ""
        titulo_tienda_online = url_tol

    producto = {
        "nombre": nombre,
        "sku": sku,
        "sku_referencial_excel": sku_ref,
        "catalogo": "tabular_completa",
        "categoria": limpiar_celda(row.get("vertical")),
        "vertical": limpiar_celda(row.get("vertical")),
        "marca": limpiar_celda(row.get("marca")) or None,
        "descripcion": descripcion,
        "indicaciones": split_indicaciones(aplicaciones),
        "aplicaciones_terapias": aplicaciones,
        "especificaciones_tecnicas": especificaciones,
        "precio_referencia_neto": limpiar_precio(row.get("precio_referencia_neto")),
        "moneda": "CLP",
        "escalable": limpiar_celda(row.get("escalable")),
        "pais_origen": limpiar_celda(row.get("pais_origen")),
        "proveedor": limpiar_celda(row.get("proveedor")),
        "perfil_comprador_ideal": limpiar_celda(row.get("perfil_comprador_ideal")),
        "tier": limpiar_celda(row.get("tier")),
        "url_web": limpiar_celda(row.get("url_web")),
        "canal_tienda_online": limpiar_celda(row.get("canal_tienda_online")),
        "url_tienda_online": url_tienda_online,
        "titulo_tienda_online": titulo_tienda_online,
        "stock": "consultar",
        "_fuente": "excel",
        "_archivo": archivo,
        "_fila_excel": fila_excel,
    }

    return producto


def procesar_excel_productos(xlsx_path: Path) -> list[dict]:
    """
    Lee el Excel tabular de Dr's Choice y lo transforma en productos estructurados.
    """
    print(f"  📊 Procesando Excel {xlsx_path.name}...")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row, headers = detectar_fila_header_excel(ws)

    productos = []

    for fila_excel, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1
    ):
        row = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[i] if i < len(values) else ""

        producto = producto_desde_excel(row, fila_excel, xlsx_path.name)
        if producto:
            productos.append(producto)

    print(f"    ✅ {len(productos)} productos desde Excel")
    return productos


def merge_productos_excel(productos_base: list[dict], productos_excel: list[dict]) -> tuple[list[dict], int, int]:
    """
    Fusiona productos del Excel con los productos existentes.
    - Si el nombre ya existe, enriquece el producto existente con precio, URL, tier, specs, etc.
    - Si no existe, lo agrega como producto nuevo.
    """
    productos = list(productos_base)

    por_nombre = {
        normalizar_clave(p.get("nombre", "")): p
        for p in productos
        if p.get("nombre")
    }

    campos_extra_excel = [
        "sku_referencial_excel",
        "precio_referencia_neto",
        "moneda",
        "vertical",
        "aplicaciones_terapias",
        "especificaciones_tecnicas",
        "escalable",
        "pais_origen",
        "proveedor",
        "perfil_comprador_ideal",
        "tier",
        "url_web",
        "canal_tienda_online",
        "url_tienda_online",
        "titulo_tienda_online",
        "_fuente",
        "_archivo",
        "_fila_excel",
    ]

    actualizados = 0
    agregados = 0

    for prod_excel in productos_excel:
        key = normalizar_clave(prod_excel.get("nombre", ""))

        if key in por_nombre:
            prod_base = por_nombre[key]

            # No pisamos datos curados importantes salvo que estén vacíos.
            for campo in ["marca", "descripcion", "categoria", "catalogo", "stock"]:
                if not prod_base.get(campo) and prod_excel.get(campo):
                    prod_base[campo] = prod_excel[campo]

            # Sí agregamos/enriquecemos campos tabulares.
            for campo in campos_extra_excel:
                valor = prod_excel.get(campo)
                if valor not in (None, "", []):
                    prod_base[campo] = valor

            # Sumamos indicaciones sin duplicar.
            indicaciones_base = prod_base.get("indicaciones", []) or []
            indicaciones_excel = prod_excel.get("indicaciones", []) or []
            prod_base["indicaciones"] = list(dict.fromkeys(indicaciones_base + indicaciones_excel))

            actualizados += 1
        else:
            productos.append(prod_excel)
            por_nombre[key] = prod_excel
            agregados += 1

    return productos, actualizados, agregados

# ── 4. Web Scraping ───────────────────────────────────────────

def cargar_urls() -> list[str]:
    """Lee las URLs desde urls.txt — una por línea, # para comentarios."""
    if not URLS_FILE.exists():
        print(f"⚠️  {URLS_FILE} no encontrado — sin web scraping")
        return []
    
    urls = []
    with open(URLS_FILE, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if linea and not linea.startswith("#"):
                urls.append(linea)
    
    print(f"📋 {len(urls)} URLs cargadas desde {URLS_FILE}")
    return urls


def scrape_url(url: str, timeout: int = 10) -> dict:
    """Extrae texto limpio de una URL."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; DrChoice-KB/2.0)"}
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        
        # Eliminamos ruido
        for tag in soup(["script", "style", "nav", "footer", "header", "meta"]):
            tag.decompose()
        
        texto = soup.get_text(separator=" ", strip=True)
        texto = re.sub(r"\s{2,}", " ", texto).strip()
        
        print(f"  🌐 {url} → {len(texto)} chars")
        return {
            "url": url,
            "texto": texto[:3000],  # máximo 3000 chars por URL
            "ok": True
        }
    except Exception as e:
        print(f"  ⚠️  {url} → Error: {e}")
        return {"url": url, "texto": "", "ok": False}


# ── 5. Ensamblado del KB ──────────────────────────────────────

def construir_knowledge_base() -> dict:
    """
    Proceso completo:
    1. Carga datos base (brandbook, empresa)
    2. Procesa todos los PDFs en /pdfs/
    3. Scrapea todas las URLs en urls.txt
    4. Ensambla el knowledge_base.json final
    """
    print("\n" + "="*60)
    print("CONSTRUYENDO KNOWLEDGE BASE — Dr's Choice")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)

    # ── Paso 1: datos base ────────────────────────────────────
    kb = cargar_base()
    kb["_generado_en"] = datetime.now().isoformat()
    kb["_fuentes"] = []

    # ── Paso 2: procesar PDFs ─────────────────────────────────
    productos_extraidos = []
    chunks_pdf = []

    if PDFS_DIR.exists():
        pdfs = sorted(PDFS_DIR.glob("*.pdf"))
        print(f"\n📂 PDFs encontrados: {len(pdfs)}")
        
        for pdf_path in pdfs:
            resultado = procesar_pdf(pdf_path)
            kb["_fuentes"].append({
                "tipo": "pdf",
                "archivo": pdf_path.name,
                "paginas": resultado["paginas"],
                "catalogo": resultado["catalogo"]
            })
            
            # Guardamos el texto completo como chunk para RAG
            chunks_pdf.append({
                "fuente": pdf_path.name,
                "catalogo": resultado["catalogo"],
                "texto": resultado["texto_completo"][:5000],
                "tablas": resultado["tablas"]
            })
            
            # Intentamos extraer productos estructurados
            prods = extraer_productos_de_texto(
                resultado["texto_completo"],
                resultado["catalogo"]
            )
            if prods:
                productos_extraidos.extend(prods)
                print(f"    → {len(prods)} productos detectados automáticamente")
    else:
        print(f"\n⚠️  Carpeta {PDFS_DIR} no encontrada — sin procesamiento de PDFs")

    # ── Paso 3: web scraping ──────────────────────────────────
    urls = cargar_urls()
    chunks_web = []
    
    if urls:
        print(f"\n🌐 Scrapeando {len(urls)} URLs...")
        for url in urls:
            resultado = scrape_url(url)
            if resultado["ok"] and resultado["texto"]:
                chunks_web.append(resultado)
                kb["_fuentes"].append({"tipo": "web", "url": url})

    # ── Paso 4: Excel tabular ─────────────────────────────────
    productos_excel = []

    if DATA_DIR.exists():
        excels = sorted(DATA_DIR.glob("*.xlsx"))
        print(f"\n📊 Excels encontrados: {len(excels)}")

        for xlsx_path in excels:
            prods_excel = procesar_excel_productos(xlsx_path)
            productos_excel.extend(prods_excel)
            kb["_fuentes"].append({
                "tipo": "excel",
                "archivo": xlsx_path.name,
                "productos": len(prods_excel)
            })
    else:
        print(f"\n⚠️  Carpeta {DATA_DIR} no encontrada — sin Excel tabular")

    # ── Paso 5: ensamblado ────────────────────────────────────
    # Productos: combinamos los existentes en la base + los extraídos de PDFs
    # Los de la base tienen prioridad (son curados manualmente)
    productos_base = kb.get("productos", [])

    skus_base = {
        str(p.get("sku", "")).strip()
        for p in productos_base
        if str(p.get("sku", "")).strip()
    }

    # Productos detectados automáticamente en PDFs.
    productos_nuevos_pdf = [
        p for p in productos_extraidos
        if str(p.get("sku", "")).strip()
        and str(p.get("sku", "")).strip() not in skus_base
    ]

    if productos_nuevos_pdf:
        print(f"\n✨ {len(productos_nuevos_pdf)} productos nuevos detectados en PDFs")

    productos_pre_excel = productos_base + productos_nuevos_pdf

    # Excel: enriquece productos existentes por nombre y agrega los nuevos.
    productos_finales, excel_actualizados, excel_agregados = merge_productos_excel(
        productos_pre_excel,
        productos_excel
    )

    kb["productos"] = productos_finales
    kb["_chunks_pdf"] = chunks_pdf
    kb["_chunks_web"] = chunks_web
    kb["_chunks_excel"] = [{
        "archivo": f.get("archivo"),
        "productos": f.get("productos")
    } for f in kb["_fuentes"] if f.get("tipo") == "excel"]
    
    print(f"\n📊 Resumen:")
    print(f"   Productos finales: {len(kb['productos'])}")
    print(f"   Productos base: {len(productos_base)}")
    print(f"   Productos nuevos desde PDF: {len(productos_nuevos_pdf)}")
    print(f"   Productos enriquecidos desde Excel: {excel_actualizados}")
    print(f"   Productos nuevos desde Excel: {excel_agregados}")
    print(f"   Chunks PDF: {len(chunks_pdf)}")
    print(f"   Chunks web: {len(chunks_web)}")
    print(f"   Fuentes totales: {len(kb['_fuentes'])}")
    
    return kb


# ── 6. Main ───────────────────────────────────────────────────

if __name__ == "__main__":
    kb = construir_knowledge_base()
    
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ {OUTPUT_FILE} generado exitosamente")
    print(f"   Tamaño: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")
